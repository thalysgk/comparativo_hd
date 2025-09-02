# analysis.py (versão com mesclagem de células no Excel)

import pandas as pd
from tkinter import messagebox
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# A função analisar_dados permanece a mesma da última versão
def analisar_dados(dados_produtos):
    if not dados_produtos:
        return pd.DataFrame()
        
    df = pd.DataFrame(dados_produtos)
    df['Preço'] = pd.to_numeric(df['Preço'], errors='coerce')
    df.dropna(subset=['Preço'], inplace=True)
    if df.empty:
        return pd.DataFrame()

    # 1. Encontra os índices (as linhas) da melhor opção para TODOS os produtos de uma vez
    idx_melhores_opcoes = df.groupby('Produto')['Preço'].idxmin()

    # 2. Cria as novas colunas, inicialmente vazias
    df['Mínimo'] = pd.NA
    df['Melhor Fornecedor'] = ''

    # 3. Usa os índices encontrados para preencher os valores apenas nas linhas corretas
    df.loc[idx_melhores_opcoes, 'Mínimo'] = df.loc[idx_melhores_opcoes, 'Preço']
    df.loc[idx_melhores_opcoes, 'Melhor Fornecedor'] = df.loc[idx_melhores_opcoes, 'Fornecedor']
    
    # Reordena as colunas para o formato final
    colunas_ordenadas = ['Produto', 'Quantidade', 'Fornecedor', 'Preço', 'Entrega', 'Mínimo', 'Melhor Fornecedor']
    # Garante que apenas colunas existentes sejam selecionadas para evitar erros
    colunas_existentes = [col for col in colunas_ordenadas if col in df.columns]
    df = df[colunas_existentes]

    return df


def _merge_product_cells(nome_arquivo):
    """
    Função auxiliar para mesclar as células da coluna 'Produto' no arquivo Excel.
    Esta é uma operação de formatação visual.
    """
    workbook = load_workbook(nome_arquivo)
    sheet = workbook.active
    
    # Dicionário para rastrear os produtos e suas linhas
    product_rows = {}
    # Começamos da linha 2 para pular o cabeçalho
    for row_idx, row in enumerate(sheet.iter_rows(min_row=2, max_col=1, values_only=True), start=2):
        product_name = row[0]
        if product_name: # Ignora células vazias se houver
            if product_name not in product_rows:
                product_rows[product_name] = []
            product_rows[product_name].append(row_idx)

    # Itera sobre os produtos e mescla as células correspondentes
    for product, rows in product_rows.items():
        if len(rows) > 1:
            start_row = min(rows)
            end_row = max(rows)
            
            # --- ALTERAÇÃO AQUI ---
            # Mescla a coluna A (Produto)
            sheet.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)
            # Mescla a coluna B (Quantidade)
            sheet.merge_cells(start_row=start_row, start_column=2, end_row=end_row, end_column=2) 
            
            # Garante que o texto fique alinhado ao centro verticalmente em ambas
            from openpyxl.styles import Alignment
            alignment = Alignment(vertical='center', horizontal='left')
            sheet.cell(row=start_row, column=1).alignment = alignment
            sheet.cell(row=start_row, column=2).alignment = alignment
            # --- FIM DA ALTERAÇÃO ---

    workbook.save(nome_arquivo)


def salvar_excel(df_final, nome_arquivo): # <-- 1. ADICIONA O PARÂMETRO 'nome_arquivo'
    """
    Salva o DataFrame final, ajusta colunas e mescla células de produto.
    """
    if df_final.empty:
        messagebox.showwarning("Aviso", "Não há dados para salvar.")
        return

    # nome_arquivo = "comparativo_final_mesclado.xlsx" # <-- 2. REMOVA ESTA LINHA FIXA
    
    try:
        df_final_sorted = df_final.sort_values(by='Produto').reset_index(drop=True)
        df_final_sorted.to_excel(nome_arquivo, sheet_name='Comparativo', index=False)
        _merge_product_cells(nome_arquivo)
        
        workbook = load_workbook(nome_arquivo)
        sheet = workbook.active
        for column_cells in sheet.columns:
            max_length = len(str(column_cells[0].value))
            for cell in column_cells:
                try:
                    if not sheet.cell(row=cell.row, column=cell.column).coordinate in sheet.merged_cells:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = adjusted_width
        workbook.save(nome_arquivo)
        
        messagebox.showinfo("Sucesso", f"Arquivo '{nome_arquivo}' salvo com sucesso!")
    except Exception as e:
        messagebox.showerror("Erro ao Salvar", f"Não foi possível salvar ou formatar o arquivo Excel.\nErro: {e}")